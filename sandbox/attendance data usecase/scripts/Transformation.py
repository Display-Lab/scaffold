"""Transform the revised obfuscated seminar workbook into the registry format.

The source workbook contains attendance sheets with date columns that represent
Fall 2025, Winter 2026, and Fall 2026 meetings. This script converts that
structure into the registry workbook while preserving the existing workbook
structure, sheet names, IDs, and formatting.
"""

from __future__ import annotations

import argparse
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_SOURCE = "source_attendance.xlsx"
DEFAULT_TEMPLATE = "registry_template.xlsx"
DEFAULT_OUTPUT = "registry_output.xlsx"

STUDENT_SHEETS = ("Students - Wn2026", "Students - Fa2025", "Students - Fa2026")
STAFF_SHEETS = ("Staff and Faculty",)


def normalize_id(value: object) -> int | str:
    if value is None:
        raise ValueError("Encountered a missing person or meeting identifier.")
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (int, str)):
        return value
    return int(value) if hasattr(value, "__int__") else str(value)


def normalize_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def normalize_lhs_701_code(value: object) -> int:
    if value is None:
        return 3

    if isinstance(value, (int, float)):
        if value == 1:
            return 1
        if value == 2:
            return 2
        return 3

    text = str(value).strip()
    if not text:
        return 3

    lowered = text.lower()
    if lowered in {"y", "yes", "yes/", "enrolled", "1", "true"}:
        return 1
    if lowered in {"n", "no", "not enrolled", "2", "false", "0"}:
        return 2
    if lowered.startswith("y"):
        return 1
    if lowered.startswith("n"):
        return 2
    return 3


def normalize_attendance_value(value: object) -> str:
    if value is None:
        return "not_attended"

    text = str(value).strip()
    if not text:
        return "not_attended"

    lowered = text.lower()
    if lowered in {"y", "yes", "in_person", "in-person", "in person"}:
        return "in_person"
    if lowered in {"y*", "remote", "remotely", "virtual", "online"}:
        return "remote"
    if lowered in {"n", "no", "not_attended", "not attended"}:
        return "not_attended"
    if lowered.startswith("y"):
        return "remote" if "*" in lowered else "in_person"
    if lowered.startswith("n"):
        return "not_attended"
    return "not_attended"


def classify_source_sheets(source_workbook) -> tuple[list[str], list[str]]:
    student_sheets: list[str] = []
    staff_sheets: list[str] = []

    for sheet_name in source_workbook.sheetnames:
        lowered = sheet_name.strip().lower()
        if lowered.startswith("students"):
            student_sheets.append(sheet_name)
            continue
        if "staff" in lowered and "faculty" in lowered:
            staff_sheets.append(sheet_name)

    if not student_sheets:
        student_sheets = [name for name in STUDENT_SHEETS if name in source_workbook.sheetnames]
    if not staff_sheets:
        staff_sheets = [name for name in STAFF_SHEETS if name in source_workbook.sheetnames]

    return student_sheets, staff_sheets


def infer_term(sheet_name: str) -> str | None:
    lowered = sheet_name.lower()
    if "fa2026" in lowered or "fall 2026" in lowered:
        return "Fall 2026"
    if "wn2026" in lowered or "winter 2026" in lowered:
        return "Winter 2026"
    if "fa2025" in lowered or "fall 2025" in lowered:
        return "Fall 2025"
    return None


def find_enrollment_column(worksheet) -> int | None:
    for index, cell in enumerate(worksheet[1], start=1):
        value = cell.value
        if isinstance(value, str) and "reg" in value.lower() and "701" in value.lower():
            return index
    return None


def find_meeting_date_columns(worksheet, start_index: int) -> list[int]:
    columns: list[int] = []
    for column_index in range(start_index, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column_index).value
        if isinstance(value, datetime) or isinstance(value, date):
            columns.append(column_index)
    return columns


def is_guest_row(worksheet, row_index: int) -> bool:
    role_value = worksheet.cell(row=row_index, column=2).value
    if role_value is None:
        return False
    lowered = str(role_value).strip().lower()
    return lowered in {"guest", "guests", "visitor", "visitors"}


def collect_meeting_dates(source_workbook, student_sheets: Iterable[str], staff_sheets: Iterable[str]) -> list[date]:
    dates = OrderedDict()
    student_set = set(student_sheets)
    for sheet_name in (*student_sheets, *staff_sheets):
        if sheet_name not in source_workbook.sheetnames:
            continue
        worksheet = source_workbook[sheet_name]
        start_index = 3 if sheet_name in student_set else 2
        for column_index in find_meeting_date_columns(worksheet, start_index):
            raw_value = worksheet.cell(row=1, column=column_index).value
            normalized = normalize_date(raw_value)
            if normalized is None:
                continue
            dates[normalized] = infer_term(sheet_name)
    return sorted(dates.keys())


def collect_people(source_workbook, student_sheets: Iterable[str], staff_sheets: Iterable[str]) -> list[dict[str, object]]:
    people = OrderedDict()

    for sheet_name in student_sheets:
        if sheet_name not in source_workbook.sheetnames:
            continue
        worksheet = source_workbook[sheet_name]
        enrollment_column = find_enrollment_column(worksheet)
        for row_index in range(2, worksheet.max_row + 1):
            person_id = worksheet.cell(row=row_index, column=1).value
            if person_id is None:
                continue
            if is_guest_row(worksheet, row_index):
                continue
            normalized_id = normalize_id(person_id)
            enrollment_value = worksheet.cell(row=row_index, column=enrollment_column).value if enrollment_column is not None else None
            existing = people.get(normalized_id)
            if existing is None:
                people[normalized_id] = {
                    "id": normalized_id,
                    "role": "Student",
                    "lhs_701_enrollment": normalize_lhs_701_code(enrollment_value),
                }
            else:
                existing["lhs_701_enrollment"] = max(existing["lhs_701_enrollment"], normalize_lhs_701_code(enrollment_value))

    for sheet_name in staff_sheets:
        if sheet_name not in source_workbook.sheetnames:
            continue
        worksheet = source_workbook[sheet_name]
        for row_index in range(2, worksheet.max_row + 1):
            person_id = worksheet.cell(row=row_index, column=1).value
            if person_id is None:
                continue
            if is_guest_row(worksheet, row_index):
                continue
            normalized_id = normalize_id(person_id)
            role_value = worksheet.cell(row=row_index, column=2).value
            role_text = str(role_value).strip() if role_value is not None else "Staff / Faculty"
            if role_text.lower() in {"faculty", "staff", "staff / faculty"}:
                role = role_text.title()
            elif "faculty" in role_text.lower():
                role = "Faculty"
            elif "staff" in role_text.lower():
                role = "Staff"
            else:
                role = "Staff / Faculty"
            people.setdefault(
                normalized_id,
                {
                    "id": normalized_id,
                    "role": role,
                    "lhs_701_enrollment": 3,
                },
            )

    return list(people.values())


def collect_attendance_events(source_workbook, student_sheets: Iterable[str], staff_sheets: Iterable[str], meeting_lookup: dict[date, int]) -> tuple[list[tuple[int, int, str]], set[tuple[int, int]], list[tuple[int, str, str]], int]:
    events: list[tuple[int, int, str]] = []
    expected_pairs: set[tuple[int, int]] = set()
    unclassified_rows: list[tuple[int, str, str]] = []
    not_attended_count = 0

    student_set = set(student_sheets)
    for sheet_name in (*student_sheets, *staff_sheets):
        if sheet_name not in source_workbook.sheetnames:
            continue
        worksheet = source_workbook[sheet_name]
        start_index = 3 if sheet_name in student_set else 2
        meeting_date_columns = find_meeting_date_columns(worksheet, start_index)
        for row_index in range(2, worksheet.max_row + 1):
            person_id = worksheet.cell(row=row_index, column=1).value
            if person_id is None or is_guest_row(worksheet, row_index):
                continue
            normalized_person_id = normalize_id(person_id)
            for column_index in meeting_date_columns:
                meeting_date = normalize_date(worksheet.cell(row=1, column=column_index).value)
                if meeting_date is None:
                    continue
                meeting_id = meeting_lookup.get(meeting_date)
                if meeting_id is None:
                    continue
                expected_pairs.add((normalized_person_id, meeting_id))
                raw_value = worksheet.cell(row=row_index, column=column_index).value
                attendance_value = normalize_attendance_value(raw_value)
                if attendance_value == "not_attended":
                    not_attended_count += 1
                if attendance_value == "not_attended" and raw_value is not None and str(raw_value).strip() not in {"", "N", "n"}:
                    unclassified_rows.append((normalized_person_id, sheet_name, str(raw_value)))
                events.append((normalized_person_id, meeting_id, attendance_value))

    unique_events: list[tuple[int, int, str]] = []
    seen_pairs: set[tuple[int, int]] = set()
    for person_id, meeting_id, attendance_value in events:
        key = (person_id, meeting_id)
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        unique_events.append((person_id, meeting_id, attendance_value))
    return unique_events, expected_pairs, unclassified_rows, not_attended_count


def clear_data_rows(worksheet) -> None:
    max_row = worksheet.max_row
    if max_row < 2:
        return
    for row_index in range(2, max_row + 1):
        for column_index in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if cell.coordinate in worksheet.merged_cells:
                continue
            cell.value = None


def read_existing_meetings(worksheet) -> list[dict[str, object]]:
    meetings: list[dict[str, object]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        meetings.append({"id": row[0], "location": row[1], "date": row[2]})
    return meetings


def read_existing_people(worksheet) -> list[dict[str, object]]:
    people: list[dict[str, object]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        people.append({"id": row[0], "role": row[1], "lhs_701_enrollment": row[2]})
    return people


def write_meetings(worksheet, meeting_dates: Iterable[date], existing_meetings: list[dict[str, object]]) -> tuple[dict[date, int], int]:
    clear_data_rows(worksheet)
    lookup: dict[date, int] = {}
    existing_by_date = {normalize_date(meeting.get("date")): meeting for meeting in existing_meetings if normalize_date(meeting.get("date")) is not None}
    used_ids = {int(meeting["id"]) for meeting in existing_meetings if meeting.get("id") is not None and isinstance(meeting.get("id"), (int, float)) and not isinstance(meeting.get("id"), bool)}
    next_id = max(used_ids, default=0) + 1

    row_index = 2
    for meeting_date in meeting_dates:
        existing_meeting = existing_by_date.get(meeting_date)
        if existing_meeting is not None:
            meeting_id = int(existing_meeting["id"])
        else:
            meeting_id = next_id
            next_id += 1
        worksheet.cell(row=row_index, column=1, value=meeting_id)
        worksheet.cell(row=row_index, column=2, value="DLHS")
        worksheet.cell(row=row_index, column=3, value=meeting_date)
        lookup[meeting_date] = meeting_id
        row_index += 1

    corrected_count = 0
    for meeting_row in existing_meetings:
        existing_date = normalize_date(meeting_row.get("date"))
        if existing_date is not None and existing_date not in lookup:
            corrected_count += 1
    return lookup, corrected_count


def write_people(worksheet, people: Iterable[dict[str, object]], existing_people: list[dict[str, object]]) -> tuple[int, list[tuple[int, int]], int]:
    clear_data_rows(worksheet)
    people_by_id = {normalize_id(person.get("id")): person for person in existing_people if person.get("id") is not None}
    converted_count = 0
    unclassified: list[tuple[int, int]] = []
    row_index = 2
    for person in people:
        person_id = normalize_id(person["id"])
        existing_person = people_by_id.get(person_id)
        if existing_person is None:
            role_value = person.get("role")
            lhs_code = int(person.get("lhs_701_enrollment", 3))
        else:
            role_value = existing_person.get("role") or person.get("role")
            lhs_code = int(person.get("lhs_701_enrollment", existing_person.get("lhs_701_enrollment", 3)))
        if lhs_code not in {1, 2, 3}:
            lhs_code = 3
            unclassified.append((int(person_id), 3))
        else:
            converted_count += 1
        worksheet.cell(row=row_index, column=1, value=person_id)
        worksheet.cell(row=row_index, column=2, value=role_value)
        worksheet.cell(row=row_index, column=3, value=lhs_code)
        row_index += 1
    return converted_count, unclassified, row_index - 2


def write_attendance_events(worksheet, events: Iterable[tuple[int, int, str]]) -> None:
    clear_data_rows(worksheet)
    row_index = 2
    for person_id, meeting_id, attendance_value in events:
        worksheet.cell(row=row_index, column=1, value=person_id)
        worksheet.cell(row=row_index, column=2, value=meeting_id)
        worksheet.cell(row=row_index, column=3, value=attendance_value)
        row_index += 1


def validate_rows(people: list[dict[str, object]], meeting_lookup: dict[date, int], attendance_events: list[tuple[int, int, str]], expected_pairs: set[tuple[int, int]], unclassified_rows: list[tuple[int, str, str]]) -> None:
    people_ids = {normalize_id(person["id"]) for person in people}
    meeting_ids = set(meeting_lookup.values())
    for person_id, meeting_id, _ in attendance_events:
        if person_id not in people_ids:
            raise ValueError(f"Invalid person_id in attendance events: {person_id}")
        if meeting_id not in meeting_ids:
            raise ValueError(f"Invalid meeting_id in attendance events: {meeting_id}")

    actual_pairs = {(person_id, meeting_id) for person_id, meeting_id, _ in attendance_events}
    missing_pairs = sorted(expected_pairs - actual_pairs)
    if missing_pairs:
        raise ValueError(f"Missing attendance events for expected pairs: {missing_pairs}")
    if len(actual_pairs) != len(attendance_events):
        raise ValueError("Duplicate attendance events detected.")
    if any(code not in {1, 2, 3} for code in [int(person.get("lhs_701_enrollment", 3)) for person in people]):
        raise ValueError("LHS 701 values must be 1, 2, or 3.")
    if not meeting_lookup:
        raise ValueError("No meetings were extracted from the source workbook.")
    if unclassified_rows:
        print(f"Unclassified rows: {unclassified_rows[:10]}")


def transform(source_path: Path, template_path: Path, output_path: Path) -> tuple[Path, dict[str, int]]:
    source_workbook = load_workbook(source_path, data_only=False)
    target_workbook = load_workbook(template_path, data_only=False)

    if "Meetings" not in target_workbook.sheetnames:
        target_workbook.create_sheet("Meetings")
    if "People" not in target_workbook.sheetnames:
        target_workbook.create_sheet("People")
    if "Attendance Events" not in target_workbook.sheetnames:
        target_workbook.create_sheet("Attendance Events")

    meetings_sheet = target_workbook["Meetings"]
    people_sheet = target_workbook["People"]
    attendance_sheet = target_workbook["Attendance Events"]

    existing_meetings = read_existing_meetings(meetings_sheet)
    existing_people = read_existing_people(people_sheet)

    student_sheets, staff_sheets = classify_source_sheets(source_workbook)

    meetings = collect_meeting_dates(source_workbook, student_sheets, staff_sheets)
    meeting_lookup, corrected_meeting_count = write_meetings(meetings_sheet, meetings, existing_meetings)

    people = collect_people(source_workbook, student_sheets, staff_sheets)
    converted_count, unclassified_rows, _ = write_people(people_sheet, people, existing_people)

    attendance_events, expected_pairs, unclassified_event_rows, not_attended_count = collect_attendance_events(source_workbook, student_sheets, staff_sheets, meeting_lookup)
    write_attendance_events(attendance_sheet, attendance_events)

    validate_rows(people, meeting_lookup, attendance_events, expected_pairs, unclassified_event_rows)

    target_workbook.save(output_path)
    stats = {
        "corrected_meetings": corrected_meeting_count,
        "not_attended_events": not_attended_count,
        "lhs_701_converted": converted_count,
        "unclassified_rows": len(unclassified_rows),
    }
    return output_path, stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", default=DEFAULT_SOURCE, help="Path to the revised obfuscated source workbook.")
    parser.add_argument("--template", default=DEFAULT_TEMPLATE, help="Path to the registry workbook template.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Path for the updated registry workbook.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_path = Path(args.source).expanduser()
    template_path = Path(args.template).expanduser()
    output_path = Path(args.output).expanduser()

    if not source_path.exists():
        raise FileNotFoundError(f"Source workbook not found: {source_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")

    output_path, stats = transform(source_path, template_path, output_path)
    print(f"Wrote updated registry workbook to {output_path}")
    print(f"Meeting dates corrected: {stats['corrected_meetings']}")
    print(f"Not-attended events added: {stats['not_attended_events']}")
    print(f"LHS 701 values converted: {stats['lhs_701_converted']}")
    print(f"Unclassified rows: {stats['unclassified_rows']}")


if __name__ == "__main__":
    main()
